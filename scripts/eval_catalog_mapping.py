"""수집된 KOSIS 카탈로그로 E2E 골든셋의 표 후보 Hit@K를 측정한다."""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from clafact.eval.table_mapping import evaluate_table_mapping
from clafact.kosis_catalog import CatalogIndex
from clafact.pipeline.query_gen import make_query


def golden_rows(workbook: Path) -> list[dict]:
    sheet = load_workbook(workbook, read_only=True, data_only=True)["E2E_추가20_검토"]
    headers = {sheet.cell(1, col).value: col for col in range(1, sheet.max_column + 1)}
    rows = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, headers["검토 상태"]).value != "완료":
            continue
        raw_table_ids = str(sheet.cell(row, headers["최종 통계표 ID"]).value or "")
        table_ids = [table_id.strip() for table_id in raw_table_ids.split("+") if table_id.strip()]
        if table_ids:
            rows.append({
                "candidate_id": sheet.cell(row, headers["candidate_id"]).value,
                "sentence": sheet.cell(row, headers["문장"]).value or "",
                "gold_table_ids": table_ids,
            })
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description="KOSIS 카탈로그 표 후보 평가")
    parser.add_argument("--catalog", default="data/catalog/kosis_mt_ztitled.json")
    parser.add_argument("--golden", default=r"..\ClaFact_완성형_핵심작업기록\ClaFact_골든셋_E2E확장_20건_검토본.xlsx")
    parser.add_argument("--out", default="reports/catalog_mapping_latest.json")
    args = parser.parse_args()
    catalog = ROOT / args.catalog
    index = CatalogIndex(catalog)
    if not index.tables:
        print("카탈로그가 비어 있습니다. sync_kosis_catalog.py로 표 메타를 먼저 수집하세요.")
        return 2
    rows = golden_rows((ROOT / args.golden).resolve())
    result = evaluate_table_mapping(
        rows, lambda sentence, top_k: [hit.__dict__ | {"TBL_ID": hit.tbl_id} for hit in index.search(make_query(sentence), top_k)], top_k=5,
    )
    output = ROOT / args.out
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result["summary"], ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
